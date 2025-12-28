import sys
import os
import asyncio
import requests
from datetime import datetime
from PIL import Image as PILImage
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QTextEdit, QTableWidget, QTableWidgetItem,
                             QProgressBar, QFileDialog, QMessageBox)
from PyQt6.QtCore import QThread, pyqtSignal, Qt
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage


class CrawlerThread(QThread):
    progress = pyqtSignal(str)
    result = pyqtSignal(list)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, url):
        super().__init__()
        self.url = url
        self.products = []
        self.is_running = True

    def stop(self):
        """크롤링 중지"""
        self.is_running = False
        self.progress.emit("사용자가 크롤링을 중지했습니다...")

    def run(self):
        try:
            asyncio.run(self.crawl())
        except Exception as e:
            if self.is_running:
                self.error.emit(f"크롤링 오류: {str(e)}")
        finally:
            self.finished.emit()

    async def crawl(self):
        if not self.is_running:
            return

        self.progress.emit("브라우저 초기화 중...")

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            )
            page = await context.new_page()

            try:
                if not self.is_running:
                    await browser.close()
                    return

                self.progress.emit(f"페이지 로딩 중: {self.url}")
                await page.goto(self.url, timeout=60000)

                if not self.is_running:
                    await browser.close()
                    return

                # 상품 리스트가 로드될 때까지 대기
                self.progress.emit("상품 목록 로딩 대기 중...")
                await page.wait_for_selector("//a[contains(@class, 'c-card-item__anchor')]", timeout=30000)
                await asyncio.sleep(2)

                if not self.is_running:
                    await browser.close()
                    return

                # Lazy Loading을 위한 점진적 스크롤
                self.progress.emit("페이지 스크롤 중 (이미지 로딩)...")
                await self.scroll_gradually(page)

                if not self.is_running:
                    await browser.close()
                    return

                # 상품 수집
                self.progress.emit("상품 정보 수집 중...")
                products = await self.extract_products(page)

                if not self.is_running:
                    await browser.close()
                    return

                self.progress.emit(f"총 {len(products)}개 상품 발견")

                # 이미지 다운로드
                if products and self.is_running:
                    await self.download_images(products)

                if self.is_running:
                    self.products = products
                    self.result.emit(products)

            except Exception as e:
                if self.is_running:
                    self.error.emit(f"페이지 처리 오류: {str(e)}")
            finally:
                await browser.close()

    async def scroll_gradually(self, page):
        """점진적 스크롤로 Lazy Loading 이미지 로드 (최대 100회 제한)"""
        last_height = await page.evaluate("document.body.scrollHeight")
        scroll_pause = 0.5
        scroll_step = 800
        max_scrolls = 100
        scroll_count = 0

        current_position = 0
        while scroll_count < max_scrolls and self.is_running:
            current_position += scroll_step
            await page.evaluate(f"window.scrollTo(0, {current_position})")
            await asyncio.sleep(scroll_pause)
            scroll_count += 1

            new_height = await page.evaluate("document.body.scrollHeight")

            self.progress.emit(f"스크롤 중... ({current_position}/{new_height}px) [{scroll_count}/{max_scrolls}회]")

            if current_position >= new_height:
                if new_height == last_height:
                    break
                last_height = new_height
                current_position = 0

    async def extract_products(self, page):
        """XPath를 사용한 상품 정보 추출"""
        products = []

        if not self.is_running:
            return products

        product_links = await page.query_selector_all("//a[contains(@class, 'c-card-item__anchor')]")

        self.progress.emit(f"상품 {len(product_links)}개 발견")

        for idx, link in enumerate(product_links, 1):
            if not self.is_running:
                break

            try:
                product = {}

                product['url'] = await link.get_attribute('href')
                data_log = await link.get_attribute('data-log-body')

                name_elem = await link.query_selector("xpath=.//span[@class='sr-only']")
                if name_elem:
                    product['name'] = await name_elem.inner_text()
                else:
                    product['name'] = "N/A"

                parent = await link.evaluate_handle("el => el.parentElement")

                price_elem = await parent.query_selector("xpath=.//strong[contains(@class, 'price')]")
                if price_elem:
                    product['price'] = await price_elem.inner_text()
                else:
                    if data_log and 'last_discount_price' in data_log:
                        import json
                        try:
                            log_data = json.loads(data_log.replace('&quot;', '"'))
                            product['price'] = log_data.get('last_discount_price', 'N/A')
                        except:
                            product['price'] = "N/A"
                    else:
                        product['price'] = "N/A"

                img_elem = await parent.query_selector("xpath=.//img")
                if img_elem:
                    product['thumbnail'] = await img_elem.get_attribute('src')
                    if not product['thumbnail'] or product['thumbnail'] == '':
                        product['thumbnail'] = await img_elem.get_attribute('data-src')
                    if not product['thumbnail'] or product['thumbnail'] == '':
                        product['thumbnail'] = await img_elem.get_attribute('data-original')
                else:
                    product['thumbnail'] = "N/A"

                product['registered_date'] = datetime.now().strftime("%Y-%m-%d")

                products.append(product)
                self.progress.emit(f"상품 {idx}/{len(product_links)} 처리 완료: {product['name'][:30]}")

            except Exception as e:
                self.progress.emit(f"상품 {idx} 추출 오류: {str(e)}")
                continue

        return products

    async def download_images(self, products):
        """썸네일 이미지 다운로드"""
        img_dir = "thumbnails"
        if not os.path.exists(img_dir):
            os.makedirs(img_dir)

        for idx, product in enumerate(products, 1):
            if not self.is_running:
                break

            try:
                if product.get('thumbnail') and product['thumbnail'] != "N/A":
                    url = product['thumbnail']
                    if url.startswith('//'):
                        url = 'https:' + url
                    elif url.startswith('/'):
                        url = 'https://www.11st.co.kr' + url

                    response = requests.get(url, timeout=10)
                    if response.status_code == 200:
                        safe_name = "".join(c for c in product['name'][:50] if c.isalnum() or c in (' ', '_'))
                        filename = f"{img_dir}/{idx}_{safe_name}.jpg"

                        with open(filename, 'wb') as f:
                            f.write(response.content)

                        product['thumbnail_local'] = filename
                        self.progress.emit(f"이미지 다운로드 {idx}/{len(products)}")

            except Exception as e:
                self.progress.emit(f"이미지 다운로드 오류 ({idx}): {str(e)}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.products = []
        self.crawler = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("11번가 카테고리 크롤러")
        self.setGeometry(100, 100, 1200, 800)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # URL 입력
        url_layout = QHBoxLayout()
        url_layout.addWidget(QLabel("카테고리 URL:"))
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText(
            "https://www.11st.co.kr/page/martplus/category?dispCtgr2No=1361105&dispCtgr3No=1361108")
        url_layout.addWidget(self.url_input)
        layout.addLayout(url_layout)

        # 버튼들
        button_layout = QHBoxLayout()
        self.start_btn = QPushButton("크롤링 시작")
        self.start_btn.clicked.connect(self.start_crawling)

        self.stop_btn = QPushButton("중지")
        self.stop_btn.clicked.connect(self.stop_crawling)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("background-color: #ff4444; color: white;")

        self.export_btn = QPushButton("엑셀로 내보내기")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)

        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.stop_btn)
        button_layout.addWidget(self.export_btn)
        layout.addLayout(button_layout)

        # 진행 상황
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # 로그
        layout.addWidget(QLabel("크롤링 로그:"))
        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(150)
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        # 결과 테이블
        layout.addWidget(QLabel("크롤링 결과:"))
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(5)
        self.result_table.setHorizontalHeaderLabels([
            "순번", "상품명", "가격", "썸네일URL", "로컬이미지경로"
        ])
        self.result_table.setColumnWidth(0, 50)
        self.result_table.setColumnWidth(1, 400)
        self.result_table.setColumnWidth(2, 100)
        self.result_table.setColumnWidth(3, 200)
        self.result_table.setColumnWidth(4, 200)
        layout.addWidget(self.result_table)

    def start_crawling(self):
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "경고", "URL을 입력해주세요.")
            return

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.export_btn.setEnabled(False)
        self.log_text.clear()
        self.result_table.setRowCount(0)
        self.progress_bar.setValue(0)

        self.log_text.append(f"크롤링 시작: {url}")

        self.crawler = CrawlerThread(url)
        self.crawler.progress.connect(self.update_progress)
        self.crawler.result.connect(self.display_results)
        self.crawler.finished.connect(self.crawling_finished)
        self.crawler.error.connect(self.show_error)
        self.crawler.start()

    def stop_crawling(self):
        if self.crawler and self.crawler.isRunning():
            self.crawler.stop()
            self.log_text.append("\n⚠ 크롤링을 중지하는 중...")
            self.stop_btn.setEnabled(False)

    def update_progress(self, message):
        self.log_text.append(message)
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )

    def display_results(self, products):
        self.products = products
        self.result_table.setRowCount(len(products))

        for row, product in enumerate(products):
            self.result_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.result_table.setItem(row, 1, QTableWidgetItem(product.get('name', '')))
            self.result_table.setItem(row, 2, QTableWidgetItem(product.get('price', '')))
            self.result_table.setItem(row, 3, QTableWidgetItem(product.get('thumbnail', '')))
            self.result_table.setItem(row, 4, QTableWidgetItem(product.get('thumbnail_local', '')))

        self.progress_bar.setValue(100)
        self.log_text.append(f"\n✓ 크롤링 완료! 총 {len(products)}개 상품")

    def crawling_finished(self):
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        if self.products:
            self.export_btn.setEnabled(True)

    def show_error(self, error_msg):
        self.log_text.append(f"\n✗ 오류: {error_msg}")
        QMessageBox.critical(self, "오류", error_msg)

    def export_to_excel(self):
        if not self.products:
            QMessageBox.warning(self, "경고", "내보낼 데이터가 없습니다.")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "엑셀 파일 저장", f"11st_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "상품 목록"

                # 헤더
                headers = ["순번", "대표이미지", "상품명", "가격", "썸네일URL", "로컬이미지경로"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 열 너비 설정
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 50
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 60
                ws.column_dimensions['F'].width = 60

                # 데이터 및 이미지 삽입
                for idx, product in enumerate(self.products, start=1):
                    row_num = idx + 1

                    # 순번
                    ws.cell(row=row_num, column=1, value=idx)
                    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')

                    # 상품명
                    ws.cell(row=row_num, column=3, value=product.get('name', ''))
                    ws.cell(row=row_num, column=3).alignment = Alignment(vertical='center', wrap_text=True)

                    # 가격
                    ws.cell(row=row_num, column=4, value=product.get('price', ''))
                    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right', vertical='center')

                    # 썸네일URL
                    ws.cell(row=row_num, column=5, value=product.get('thumbnail', ''))

                    # 로컬이미지경로
                    local_path = product.get('thumbnail_local', '')
                    ws.cell(row=row_num, column=6, value=local_path)

                    # 이미지 삽입 (B열)
                    if local_path and os.path.exists(local_path):
                        try:
                            # 절대 경로로 변환
                            abs_local_path = os.path.abspath(local_path)

                            # PIL로 이미지 열어서 RGB로 변환
                            pil_img = PILImage.open(abs_local_path)
                            if pil_img.mode in ('RGBA', 'LA', 'P'):
                                # 흰 배경으로 변환
                                background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                                if pil_img.mode == 'P':
                                    pil_img = pil_img.convert('RGBA')
                                background.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                                pil_img = background
                            elif pil_img.mode != 'RGB':
                                pil_img = pil_img.convert('RGB')

                            # 임시 파일 경로 (절대 경로)
                            temp_dir = os.path.dirname(abs_local_path)
                            temp_filename = f"temp_{idx}_{os.path.basename(abs_local_path)}"
                            temp_path = os.path.join(temp_dir, temp_filename)

                            # 임시 파일로 저장
                            pil_img.save(temp_path, 'JPEG', quality=95)

                            # openpyxl로 이미지 삽입
                            img = XLImage(temp_path)
                            img.width = 100
                            img.height = 100

                            # B열에 이미지 추가
                            ws.add_image(img, f'B{row_num}')

                            # 행 높이 조정
                            ws.row_dimensions[row_num].height = 75

                            self.log_text.append(f"✓ 이미지 삽입: 행 {row_num}")

                            # 임시 파일 삭제 (워크북 저장 후 삭제하도록 리스트에 저장)
                            if not hasattr(self, 'temp_files'):
                                self.temp_files = []
                            self.temp_files.append(temp_path)

                        except Exception as e:
                            self.log_text.append(f"✗ 이미지 삽입 오류 (행 {row_num}): {str(e)}")
                            ws.row_dimensions[row_num].height = 30
                    else:
                        self.log_text.append(f"⚠ 이미지 파일 없음 (행 {row_num}): {local_path}")
                        ws.row_dimensions[row_num].height = 30

                # 엑셀 파일 저장
                wb.save(filename)

                # 임시 파일 삭제
                if hasattr(self, 'temp_files'):
                    for temp_file in self.temp_files:
                        try:
                            if os.path.exists(temp_file):
                                os.remove(temp_file)
                        except:
                            pass
                    self.temp_files = []

                self.log_text.append(f"\n✓ 엑셀 파일 저장 완료: {filename}")
                QMessageBox.information(self, "성공", f"엑셀 파일이 저장되었습니다.\n{filename}")

            except Exception as e:
                QMessageBox.critical(self, "오류", f"엑셀 저장 실패: {str(e)}")
                self.log_text.append(f"\n✗ 상세 오류: {str(e)}")


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)

